from __future__ import annotations

import re
import uuid
from dataclasses import dataclass
from datetime import datetime
from io import BytesIO
from typing import Any

from sqlalchemy.orm import Session

from app.core.config import settings
from app.models.product import Product, ProductStatus
from app.models.pricing_rule import PricingRule


class PriceListImportError(RuntimeError):
    pass


@dataclass
class PriceListImportResult:
    created: int = 0
    updated: int = 0
    skipped: int = 0
    pending: int = 0
    missing: int = 0
    total_seen: int = 0


class PriceListImportService:
    """One-time / repeated ACTIVE.xlsx import.

    Existing products are updated, new products are created, old products are not deleted.
    Products not found in the latest import are marked missing_from_last_import=True.
    """

    def _load_openpyxl(self):
        try:
            from openpyxl import load_workbook
        except Exception as exc:
            raise PriceListImportError('Не установлен openpyxl. В терминале выполни: pip install openpyxl') from exc
        return load_workbook

    def import_xlsx(self, db: Session, *, store_id: int, data: bytes) -> PriceListImportResult:
        load_workbook = self._load_openpyxl()
        try:
            wb = load_workbook(BytesIO(data), data_only=True)
        except Exception as exc:
            raise PriceListImportError('Не смог открыть Excel. Загрузи именно ACTIVE.xlsx / прайс-лист из Kaspi.') from exc

        ws = wb.worksheets[0]
        header_row = self._find_header_row(ws)
        columns = self._find_columns(ws, header_row)
        sku_col = self._column(columns, ['sku', 'артикул'])
        price_col = self._column(columns, ['price', 'цена'])
        model_col = self._column(columns, ['model', 'модель', 'name', 'название', 'наименование'])
        brand_col = self._column(columns, ['brand', 'бренд'])
        pp1_col = self._column(columns, ['pp1', 'остаток', 'stock', 'количество'])
        url_col = self._column(columns, ['url', 'ссылка', 'link'])

        if not sku_col or not price_col:
            raise PriceListImportError('В Excel не найдены колонки SKU и price. Не меняй названия колонок в файле Kaspi.')

        batch_id = datetime.utcnow().strftime('%Y%m%d%H%M%S_') + uuid.uuid4().hex[:6]
        now = datetime.utcnow()
        # mark current rows as missing; imported rows below will turn this off
        db.query(Product).filter(Product.store_id == store_id).update({
            Product.missing_from_last_import: True,
        }, synchronize_session=False)

        result = PriceListImportResult()
        seen: set[str] = set()
        for row in range(header_row + 1, ws.max_row + 1):
            sku = self._norm_sku(ws.cell(row=row, column=sku_col).value)
            if not sku or sku in seen:
                result.skipped += 1
                continue
            seen.add(sku)
            result.total_seen += 1
            price = self._num(ws.cell(row=row, column=price_col).value)
            if price <= 0:
                result.skipped += 1
                continue
            model = str(ws.cell(row=row, column=model_col).value or '').strip() if model_col else ''
            name = model or sku
            brand = str(ws.cell(row=row, column=brand_col).value or '').strip() if brand_col else (settings.KASPI_DEFAULT_BRAND or '')
            stock = int(self._num(ws.cell(row=row, column=pp1_col).value)) if pp1_col else 0
            url = str(ws.cell(row=row, column=url_col).value or '').strip() if url_col else ''
            product_id = self._extract_product_id(' '.join([sku, url]))

            product = db.query(Product).filter(Product.store_id == store_id, Product.kaspi_sku == sku).first()
            if product:
                product.current_price = price
                product.name = (name or product.name or sku)[:255]
                product.model = (model or product.model or name or sku)[:255]
                if brand:
                    product.brand = brand[:120]
                if url:
                    product.url = url
                if product_id:
                    product.product_id = product_id[:120]
                if pp1_col:
                    product.stock = stock
                product.last_imported_at = now
                product.last_seen_import_batch = batch_id
                product.missing_from_last_import = False
                db.add(product)
                if not product.pricing_rule:
                    db.add(PricingRule(product_id=product.id))
                result.updated += 1
            else:
                product = Product(
                    store_id=store_id,
                    kaspi_sku=sku,
                    product_id=product_id[:120],
                    name=(name or sku)[:255],
                    model=(model or name or sku)[:255],
                    brand=brand[:120],
                    url=url,
                    current_price=price,
                    min_price=0,
                    max_price=0,
                    cost_price=0,
                    stock=stock,
                    status=ProductStatus.PAUSED,
                    auto_pricing_enabled=False,
                    last_imported_at=now,
                    last_seen_import_batch=batch_id,
                    missing_from_last_import=False,
                )
                db.add(product)
                db.flush()
                db.add(PricingRule(product_id=product.id))
                result.created += 1

        result.missing = db.query(Product).filter(Product.store_id == store_id, Product.missing_from_last_import == True).count()
        db.commit()
        result.pending = self.count_pending(db, store_id=store_id)
        return result

    def count_pending(self, db: Session, store_id: int | None = None) -> int:
        q = db.query(Product).filter(
            (Product.auto_pricing_enabled == False) | (Product.min_price <= 0) | (Product.max_price <= 0)
        )
        if store_id:
            q = q.filter(Product.store_id == store_id)
        return q.count()

    def _find_header_row(self, ws) -> int:
        for row in range(1, min(ws.max_row, 30) + 1):
            values = {str(ws.cell(row=row, column=col).value or '').strip().lower() for col in range(1, ws.max_column + 1)}
            if ('sku' in values or 'артикул' in values) and ('price' in values or 'цена' in values):
                return row
        return 1

    def _find_columns(self, ws, header_row: int) -> dict[str, int]:
        columns: dict[str, int] = {}
        for col in range(1, ws.max_column + 1):
            value = str(ws.cell(row=header_row, column=col).value or '').strip().lower()
            if value:
                columns[value] = col
        return columns

    def _column(self, columns: dict[str, int], names: list[str]) -> int | None:
        for header, col in columns.items():
            h = header.lower().strip()
            for name in names:
                n = name.lower().strip()
                if h == n or n in h:
                    return col
        return None

    def _norm_sku(self, value: Any) -> str:
        if value is None:
            return ''
        text = str(value).strip()
        if text.endswith('.0') and text[:-2].isdigit():
            text = text[:-2]
        return text

    def _extract_product_id(self, text: str) -> str:
        text = str(text or '')
        m = re.search(r'(?:-|/)(\d{6,})(?:/|$|_)', text)
        if m:
            return m.group(1)
        m = re.search(r'\d{6,}', text)
        return m.group(0) if m else ''

    def _num(self, value: Any) -> float:
        if value is None or value == '':
            return 0.0
        if isinstance(value, (int, float)):
            return float(value)
        text = str(value).strip().replace('\xa0', '').replace(' ', '').replace('₸', '').replace(',', '.')
        try:
            return float(text)
        except ValueError:
            return 0.0


price_list_import_service = PriceListImportService()
