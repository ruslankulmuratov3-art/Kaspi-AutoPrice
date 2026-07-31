from __future__ import annotations

import base64
import json
import re
import uuid
from datetime import datetime
from pathlib import Path
from typing import Any

from app.core.database import SessionLocal
from app.models.task_log import TaskLog, TaskStatus


class GeneratedPriceListService:
    """Archive generated Excel/XML files.

    Saves files locally for fast download and duplicates metadata/content to the DB.
    On Render this means records survive redeploy when DATABASE_URL points to PostgreSQL.
    """

    def __init__(self, base_dir: str = 'storage/generated_price_lists'):
        self.base_dir = Path(base_dir)
        self.index_path = self.base_dir / 'index.json'

    def ensure_dir(self) -> None:
        self.base_dir.mkdir(parents=True, exist_ok=True)

    def _now(self) -> datetime:
        return datetime.utcnow()

    def _read_index(self) -> list[dict[str, Any]]:
        self.ensure_dir()
        if not self.index_path.exists():
            return []
        try:
            data = json.loads(self.index_path.read_text(encoding='utf-8'))
            return data if isinstance(data, list) else []
        except Exception:
            return []

    def _write_index(self, rows: list[dict[str, Any]]) -> None:
        self.ensure_dir()
        self.index_path.write_text(json.dumps(rows, ensure_ascii=False, indent=2), encoding='utf-8')

    def _save_db_record(self, record: dict[str, Any], data: bytes, file_type: str = 'excel') -> None:
        try:
            db = SessionLocal()
            payload = {
                **record,
                'file_type': file_type,
                'content_base64': base64.b64encode(data).decode('ascii'),
            }
            db.add(TaskLog(
                task_name='generated_price_list',
                task_id=str(record.get('id') or uuid.uuid4().hex[:12]),
                status=TaskStatus.SUCCESS,
                message=str(record.get('filename') or ''),
                payload_json=json.dumps(payload, ensure_ascii=False),
            ))
            db.commit()
        except Exception:
            try:
                db.rollback()
            except Exception:
                pass
        finally:
            try:
                db.close()
            except Exception:
                pass

    def save_excel(
        self,
        data: bytes,
        *,
        store_id: int | None,
        store_name: str,
        source_filename: str,
        total_rows: int,
        processed_products: int,
        changed: int,
        skipped: int,
        scope: str,
        selected_count: int = 0,
        q_filter: str = '',
    ) -> dict[str, Any]:
        self.ensure_dir()
        now = self._now()
        record_id = uuid.uuid4().hex[:12]
        safe_store = self._slug(store_name or f'store_{store_id or 0}')
        filename = f'kaspi_ready_{safe_store}_{now.strftime("%Y%m%d_%H%M%S")}_{record_id}.xlsx'
        path = self.base_dir / filename
        path.write_bytes(data)
        record = {
            'id': record_id,
            'filename': filename,
            'path': str(path.as_posix()),
            'created_at': now.isoformat(timespec='seconds'),
            'store_id': store_id,
            'store_name': store_name or '',
            'source_filename': source_filename or '',
            'size_bytes': len(data),
            'size_label': self.size_label(len(data)),
            'total_rows': int(total_rows or 0),
            'processed_products': int(processed_products or 0),
            'changed': int(changed or 0),
            'skipped': int(skipped or 0),
            'scope': scope or '',
            'selected_count': int(selected_count or 0),
            'q_filter': q_filter or '',
            'type': 'Excel',
            'status': 'saved',
        }
        rows = self._read_index()
        rows.insert(0, record)
        self._write_index(rows[:300])
        self._save_db_record(record, data, 'excel')
        return record

    def _db_records(self, *, store_id: int | None = None, limit: int = 30) -> list[dict[str, Any]]:
        rows: list[dict[str, Any]] = []
        try:
            db = SessionLocal()
            logs = db.query(TaskLog).filter(TaskLog.task_name == 'generated_price_list').order_by(TaskLog.created_at.desc()).limit(max(50, limit * 5)).all()
            for log in logs:
                try:
                    payload = json.loads(log.payload_json or '{}')
                except Exception:
                    continue
                payload.pop('content_base64', None)
                if store_id and int(payload.get('store_id') or 0) != int(store_id):
                    continue
                rows.append(payload)
                if len(rows) >= limit:
                    break
        except Exception:
            rows = []
        finally:
            try:
                db.close()
            except Exception:
                pass
        return rows

    def list_records(self, *, store_id: int | None = None, limit: int = 30) -> list[dict[str, Any]]:
        rows = self._db_records(store_id=store_id, limit=limit)
        if rows:
            return rows[:limit]
        rows = self._read_index()
        if store_id:
            rows = [r for r in rows if int(r.get('store_id') or 0) == int(store_id)]
        return rows[:limit]

    def get_record(self, record_id: str) -> dict[str, Any] | None:
        try:
            db = SessionLocal()
            log = db.query(TaskLog).filter(TaskLog.task_name == 'generated_price_list', TaskLog.task_id == str(record_id)).order_by(TaskLog.id.desc()).first()
            if log:
                payload = json.loads(log.payload_json or '{}')
                payload.pop('content_base64', None)
                return payload
        except Exception:
            pass
        finally:
            try:
                db.close()
            except Exception:
                pass
        for row in self._read_index():
            if str(row.get('id')) == str(record_id) or str(row.get('filename')) == str(record_id):
                return row
        return None

    def get_file_bytes(self, record_id: str) -> bytes | None:
        # local file first
        record = self.get_record(record_id)
        if record:
            path = self.file_path_for(record)
            if path and path.exists():
                return path.read_bytes()
        # DB fallback
        try:
            db = SessionLocal()
            log = db.query(TaskLog).filter(TaskLog.task_name == 'generated_price_list', TaskLog.task_id == str(record_id)).order_by(TaskLog.id.desc()).first()
            if not log:
                return None
            payload = json.loads(log.payload_json or '{}')
            b64 = payload.get('content_base64')
            if not b64:
                return None
            return base64.b64decode(b64)
        except Exception:
            return None
        finally:
            try:
                db.close()
            except Exception:
                pass

    def file_path_for(self, record: dict[str, Any]) -> Path | None:
        filename = str(record.get('filename') or '')
        if not filename or '/' in filename or '\\' in filename or '..' in filename:
            return None
        path = (self.base_dir / filename).resolve()
        base = self.base_dir.resolve()
        if base not in path.parents and path != base:
            return None
        return path if path.exists() else None

    def count_sku_rows(self, data: bytes) -> int:
        try:
            from io import BytesIO
            from openpyxl import load_workbook
            wb = load_workbook(BytesIO(data), read_only=True, data_only=True)
            ws = wb.worksheets[0]
            header_row = 1
            sku_col = None
            for row_idx in range(1, min(ws.max_row or 1, 20) + 1):
                values = [str(ws.cell(row=row_idx, column=col).value or '').strip().lower() for col in range(1, (ws.max_column or 1) + 1)]
                if 'sku' in values or 'артикул' in values:
                    header_row = row_idx
                    sku_col = (values.index('sku') + 1) if 'sku' in values else (values.index('артикул') + 1)
                    break
            if sku_col is None:
                return 0
            return sum(1 for row_idx in range(header_row + 1, (ws.max_row or header_row) + 1) if str(ws.cell(row=row_idx, column=sku_col).value or '').strip())
        except Exception:
            return 0

    def size_label(self, n: int) -> str:
        n = int(n or 0)
        if n >= 1024 * 1024:
            return f'{n / (1024 * 1024):.1f} МБ'
        if n >= 1024:
            return f'{n / 1024:.1f} КБ'
        return f'{n} Б'

    def _slug(self, text: str) -> str:
        text = re.sub(r'[^a-zA-Z0-9_-]+', '_', text.strip())
        return text.strip('_')[:40] or 'store'


generated_price_list_service = GeneratedPriceListService()
