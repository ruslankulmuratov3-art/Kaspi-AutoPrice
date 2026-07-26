from __future__ import annotations

import json
import re
import uuid
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any


@dataclass
class GeneratedPriceListRecord:
    id: str
    filename: str
    path: str
    created_at: str
    store_id: int | None
    store_name: str
    source_filename: str
    size_bytes: int
    size_label: str
    total_rows: int
    processed_products: int
    changed: int
    skipped: int
    scope: str
    selected_count: int
    q_filter: str


class GeneratedPriceListService:
    """Simple local archive for generated Kaspi Excel files.

    No DB migration needed: metadata is stored in storage/generated_price_lists/index.json,
    files are stored рядом в этой же папке.
    """

    def __init__(self, base_dir: str = 'storage/generated_price_lists'):
        self.base_dir = Path(base_dir)
        self.index_path = self.base_dir / 'index.json'

    def ensure_dir(self) -> None:
        self.base_dir.mkdir(parents=True, exist_ok=True)

    def _read_index(self) -> list[dict[str, Any]]:
        self.ensure_dir()
        if not self.index_path.exists():
            return []
        try:
            data = json.loads(self.index_path.read_text(encoding='utf-8'))
            return data if isinstance(data, list) else []
        except Exception:
            return []

    def _write_index(self, rows: list[dict[str, Any]]) -> None:
        self.ensure_dir()
        self.index_path.write_text(json.dumps(rows, ensure_ascii=False, indent=2), encoding='utf-8')

    def save_excel(
        self,
        data: bytes,
        *,
        store_id: int | None,
        store_name: str,
        source_filename: str,
        total_rows: int,
        processed_products: int,
        changed: int,
        skipped: int,
        scope: str,
        selected_count: int = 0,
        q_filter: str = '',
    ) -> dict[str, Any]:
        self.ensure_dir()
        now = datetime.now()
        record_id = uuid.uuid4().hex[:12]
        safe_store = self._slug(store_name or f'store_{store_id or 0}')
        filename = f'kaspi_ready_{safe_store}_{now.strftime("%Y%m%d_%H%M%S")}_{record_id}.xlsx'
        path = self.base_dir / filename
        path.write_bytes(data)
        record = {
            'id': record_id,
            'filename': filename,
            'path': str(path.as_posix()),
            'created_at': now.isoformat(timespec='seconds'),
            'store_id': store_id,
            'store_name': store_name or '',
            'source_filename': source_filename or '',
            'size_bytes': len(data),
            'size_label': self.size_label(len(data)),
            'total_rows': int(total_rows or 0),
            'processed_products': int(processed_products or 0),
            'changed': int(changed or 0),
            'skipped': int(skipped or 0),
            'scope': scope or '',
            'selected_count': int(selected_count or 0),
            'q_filter': q_filter or '',
        }
        rows = self._read_index()
        rows.insert(0, record)
        rows = rows[:300]
        self._write_index(rows)
        return record

    def list_records(self, *, store_id: int | None = None, limit: int = 30) -> list[dict[str, Any]]:
        rows = self._read_index()
        if store_id:
            rows = [r for r in rows if int(r.get('store_id') or 0) == int(store_id)]
        return rows[:limit]

    def get_record(self, record_id: str) -> dict[str, Any] | None:
        for row in self._read_index():
            if str(row.get('id')) == str(record_id):
                return row
            # fallback: allow direct filename match if needed
            if str(row.get('filename')) == str(record_id):
                return row
        return None

    def file_path_for(self, record: dict[str, Any]) -> Path | None:
        filename = str(record.get('filename') or '')
        if not filename or '/' in filename or '\\' in filename or '..' in filename:
            return None
        path = self.base_dir / filename
        try:
            path = path.resolve()
            base = self.base_dir.resolve()
            if base not in path.parents and path != base:
                return None
        except Exception:
            return None
        return path if path.exists() else None

    def count_sku_rows(self, data: bytes) -> int:
        """Count SKU rows in a Kaspi Excel file. Best effort; returns 0 if cannot read."""
        try:
            from io import BytesIO
            from openpyxl import load_workbook
            wb = load_workbook(BytesIO(data), read_only=True, data_only=True)
            ws = wb.worksheets[0]
            header_row = 1
            sku_col = None
            for row_idx in range(1, min(ws.max_row or 1, 20) + 1):
                values = [str(ws.cell(row=row_idx, column=col).value or '').strip().lower() for col in range(1, (ws.max_column or 1) + 1)]
                if 'sku' in values:
                    header_row = row_idx
                    sku_col = values.index('sku') + 1
                    break
            if sku_col is None:
                return 0
            count = 0
            for row_idx in range(header_row + 1, (ws.max_row or header_row) + 1):
                value = ws.cell(row=row_idx, column=sku_col).value
                if str(value or '').strip():
                    count += 1
            return count
        except Exception:
            return 0

    def size_label(self, n: int) -> str:
        n = int(n or 0)
        if n >= 1024 * 1024:
            return f'{n / (1024 * 1024):.1f} МБ'
        if n >= 1024:
            return f'{n / 1024:.1f} КБ'
        return f'{n} Б'

    def _slug(self, text: str) -> str:
        text = re.sub(r'[^a-zA-Z0-9_-]+', '_', text.strip())
        return text.strip('_')[:40] or 'store'


generated_price_list_service = GeneratedPriceListService()
