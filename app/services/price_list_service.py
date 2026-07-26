from __future__ import annotations

from io import BytesIO
from pathlib import Path
from typing import Iterable

from app.core.config import settings
from app.models.product import Product


class PriceListError(RuntimeError):
    pass


class PriceListService:
    """Create a FULL Kaspi Excel price list without calling the product import API.

    Безопасный режим: всегда берём полный ACTIVE.xlsx как шаблон, меняем только колонку
    price у найденных SKU, а остальные строки/остатки/колонки оставляем как были.
    Минимальные файлы на один товар запрещены, чтобы не рисковать снятием товаров с продажи.
    """

    HEADERS = ['SKU', 'model', 'brand', 'price', 'PP1', 'PP2', 'PP3', 'PP4', 'PP5', 'preorder']

    def build_xlsx(
        self,
        products: Iterable[Product],
        price_by_sku: dict[str, int],
        *,
        template_bytes: bytes | None = None,
    ) -> bytes:
        if not template_bytes:
            raise PriceListError('Нужно выбрать полный ACTIVE.xlsx из Kaspi. Минимальный Excel на один товар больше не создаём, чтобы случайно не снять остальные товары с продажи.')
        return self._update_template(template_bytes, price_by_sku)

    def _load_openpyxl(self):
        try:
            from openpyxl import Workbook, load_workbook
            from openpyxl.styles import Font
        except Exception as exc:  # pragma: no cover - depends on local install
            raise PriceListError('Не установлен openpyxl. В терминале выполни: pip install openpyxl') from exc
        return Workbook, load_workbook, Font

    def _update_template(self, template_bytes: bytes, price_by_sku: dict[str, int]) -> bytes:
        _, load_workbook, _ = self._load_openpyxl()
        try:
            wb = load_workbook(BytesIO(template_bytes))
        except Exception as exc:
            raise PriceListError('Не смог открыть Excel-шаблон. Скачай прайс-лист из Kaspi в формате .xlsx и попробуй снова.') from exc
        ws = wb.worksheets[0]
        header_row = self._find_header_row(ws)
        columns = self._find_columns(ws, header_row)
        sku_col = columns.get('sku')
        price_col = columns.get('price')
        if not sku_col or not price_col:
            raise PriceListError('В шаблоне не найдены обязательные колонки SKU и price.')

        changed = 0
        normalized_prices = {self._norm_sku(sku): int(price) for sku, price in price_by_sku.items()}
        for row in range(header_row + 1, ws.max_row + 1):
            sku = self._norm_sku(ws.cell(row=row, column=sku_col).value)
            if sku and sku in normalized_prices:
                ws.cell(row=row, column=price_col).value = normalized_prices[sku]
                changed += 1
        if changed == 0:
            sample = ', '.join(list(price_by_sku.keys())[:5])
            raise PriceListError(f'Ни один SKU из сайта не найден в Excel-шаблоне Kaspi. Проверяемые SKU: {sample}')
        return self._save_workbook(wb)

    def _build_minimal(self, products: Iterable[Product], price_by_sku: dict[str, int]) -> bytes:
        Workbook, _, Font = self._load_openpyxl()
        wb = Workbook()
        ws = wb.active
        ws.title = 'price-list'
        ws.append(self.HEADERS)
        for cell in ws[1]:
            cell.font = Font(bold=True)

        rows = 0
        for product in products:
            sku = str(product.kaspi_sku or '').strip()
            if not sku:
                continue
            price = int(price_by_sku.get(sku, int(round(float(product.current_price or 0)))))
            if price <= 0:
                continue
            # В минимальном файле НЕ ставим 0/no в PP1, чтобы случайно не обнулить остатки.
            stock_value = int(product.stock or 0) if int(product.stock or 0) > 0 else ''
            ws.append([
                sku,
                product.name or sku,
                product.brand or settings.KASPI_DEFAULT_BRAND or '',
                price,
                stock_value,
                '', '', '', '',
                '',
            ])
            rows += 1
        if rows == 0:
            raise PriceListError('Нет товаров с ценой больше 0 для прайс-листа.')
        widths = {'A': 24, 'B': 48, 'C': 18, 'D': 12, 'E': 10, 'F': 10, 'G': 10, 'H': 10, 'I': 10, 'J': 12}
        for col, width in widths.items():
            ws.column_dimensions[col].width = width
        return self._save_workbook(wb)

    def _save_workbook(self, wb) -> bytes:
        output = BytesIO()
        wb.save(output)
        return output.getvalue()

    def _find_header_row(self, ws) -> int:
        for row in range(1, min(ws.max_row, 20) + 1):
            values = {str(ws.cell(row=row, column=col).value or '').strip().lower() for col in range(1, ws.max_column + 1)}
            if 'sku' in values and 'price' in values:
                return row
        return 1

    def _find_columns(self, ws, header_row: int) -> dict[str, int]:
        columns: dict[str, int] = {}
        for col in range(1, ws.max_column + 1):
            value = str(ws.cell(row=header_row, column=col).value or '').strip().lower()
            if value:
                columns[value] = col
        return columns

    def _norm_sku(self, value) -> str:
        if value is None:
            return ''
        text = str(value).strip()
        if text.endswith('.0') and text.replace('.0', '').isdigit():
            text = text[:-2]
        return text


price_list_service = PriceListService()
